import bs4
import os
import openpyxl
from urllib.request import urlopen


def yahoo_dict(word):
  html = urlopen("https://tw.dictionary.search.yahoo.com/search?p=%s" % word).read().decode("utf-8")
  bs = bs4.BeautifulSoup(html, "html.parser")

  explainArea = bs.select(".explain.DictionaryResults")[0]
  wordClassList = explainArea.select(".compTitle")
  wordExplainsList = explainArea.select(".compArticleList")

  meaning = ""
  for i in range(len(wordClassList)):
    wordClass = wordClassList[i].getText()
    wordExplains = [h4.getText() for h4 in wordExplainsList[i].select("h4")]

    meaning += "%s\n" % wordClass
    for exp in wordExplains:
      meaning += "  %s\n" % exp
  return meaning


while True:
    learn_or_review=input('你今天想學新字還是複習啊ＸＤ？ 輸入l 或 r:')
    if learn_or_review == 'l':

      FILE_NAME = 'Newwords.xlsx'

      if os.path.isfile(FILE_NAME):
        wb = openpyxl.load_workbook(FILE_NAME)
      else:
        wb = openpyxl.Workbook()
        wb.remove_sheet(wb.active)

      Chapter  = input('這是第幾章節?: ')
      sheet_name = "%02d" % (int(Chapter))

      if sheet_name in wb.get_sheet_names():
        sh = wb.get_sheet_by_name(sheet_name)
      else:
        sh = wb.create_sheet(sheet_name)

      while True:
        word = input("請輸入單字: ")
        nextRow = len(sh.rows) + 1
        sh.cell(row=nextRow, column=1).value = word
        sh.cell(row=nextRow, column=2).value = yahoo_dict(word)

        wb.save(filename=FILE_NAME)
        print(word+':')
        print(yahoo_dict(word))
        print()

    if learn_or_review == 'r':

        #複習章節
        while True:
            wb = openpyxl.load_workbook('Newwords.xlsx')
            Chapter  = input('這是第幾章節?: ')
            sheet_name = "%02d" % (int(Chapter))
            sheet_list=wb.get_sheet_names()
            
            try:
                sh = wb.get_sheet_by_name(sheet_name)
                break
            except:
                print("no sheet in Newwords.xlsx,here are your chapters:")
                for sheet_names in sheet_list:
                    print(sheet_names)
                print()
                continue

        #單字表
        word_list=[]
        for row in range(1,len(sh.rows)):
            word=sh.cell(row=row, column=1).value
            meaning=sh.cell(row=row, column=2).value
            word_content=word+':'+meaning
            word_list.append(word_content)
        #總共有幾個字
        print('you got',len(word_list),'words to memorize  QQ')

        #考試時間
        import random
        while len(word_list)>0:
            random.shuffle(word_list)
            wordandmean=word_list[0].split(':')
            word=wordandmean[0]
            meaning=wordandmean[1]
            print(word)
            answer=input('do you remember the word?? y or n:')
            if answer=='y':
                word_list.pop(0)
                print(meaning)
                print("Good job! Let's move on to next word!!")
                print("word left: ", len(word_list))
                print()
            elif answer=='n':
                print("The meaning is:\n"+meaning)
                print("SHIT!let's do it again!!")
                print("word left: ", len(word_list))
                print()
            else:
                print('please answer yes or no, damn it!')
                print("word left: ", len(word_list))
                print()
        while len(word_list)==0:
            print('Congratulations! You have memerized all the words!!')
            print()
            print()
            break

    else:
        print('叫你輸入l或r聽不懂嗎？笨豬！')
        continue
            
    
